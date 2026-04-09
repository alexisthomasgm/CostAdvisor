import io
import uuid
from collections import defaultdict
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.user import User
from app.models.supplier import Supplier
from app.models.team import TeamMembership
from app.models.cost_model import CostModel
from app.models.price_data import ActualPrice
from app.models.actual_volume import ActualVolume
from app.routers.auth import get_current_user
from app.schemas.supplier import SupplierCreate, SupplierOut
from app.services.audit import log_event

router = APIRouter()


def require_team_access(db: Session, user: User, team_id: uuid.UUID):
    membership = db.query(TeamMembership).filter(
        TeamMembership.user_id == user.id,
        TeamMembership.team_id == team_id,
    ).first()
    if not membership:
        raise HTTPException(status_code=403, detail="Not a member of this team")


@router.get("/", response_model=list[SupplierOut])
def list_suppliers(
    team_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    require_team_access(db, current_user, team_id)
    return db.query(Supplier).filter(Supplier.team_id == team_id).order_by(Supplier.name).all()


@router.post("/", response_model=SupplierOut, status_code=201)
def create_supplier(
    team_id: uuid.UUID,
    data: SupplierCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    require_team_access(db, current_user, team_id)
    supplier = Supplier(
        team_id=team_id,
        name=data.name,
        country=data.country,
    )
    db.add(supplier)
    log_event(db, team_id, current_user.id, "create", "supplier", str(supplier.id),
              new_value={"name": data.name, "country": data.country})
    db.commit()
    db.refresh(supplier)
    return supplier


@router.put("/{supplier_id}", response_model=SupplierOut)
def update_supplier(
    supplier_id: int,
    data: SupplierCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    require_team_access(db, current_user, supplier.team_id)
    prev = {"name": supplier.name, "country": supplier.country}
    supplier.name = data.name
    supplier.country = data.country
    log_event(db, supplier.team_id, current_user.id, "update", "supplier", str(supplier.id),
              previous_value=prev, new_value={"name": data.name, "country": data.country})
    db.commit()
    db.refresh(supplier)
    return supplier


@router.get("/{supplier_id}/purchase-history")
def get_purchase_history(
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    require_team_access(db, current_user, supplier.team_id)

    models = (
        db.query(CostModel)
        .filter(CostModel.supplier_id == supplier_id)
        .all()
    )
    model_ids = [m.id for m in models]

    prices = (
        db.query(ActualPrice)
        .filter(ActualPrice.cost_model_id.in_(model_ids))
        .order_by(ActualPrice.year, ActualPrice.quarter)
        .all()
    ) if model_ids else []

    volumes = (
        db.query(ActualVolume)
        .filter(ActualVolume.cost_model_id.in_(model_ids))
        .order_by(ActualVolume.year, ActualVolume.quarter)
        .all()
    ) if model_ids else []

    # Index volumes by (cost_model_id, year, quarter)
    vol_map = {}
    for v in volumes:
        vol_map[(str(v.cost_model_id), v.year, v.quarter)] = {
            "volume": float(v.volume),
            "unit": v.unit,
        }

    # Group prices by cost_model_id and merge with volumes
    price_map = defaultdict(list)
    for p in prices:
        key = (str(p.cost_model_id), p.year, p.quarter)
        vol = vol_map.get(key, {})
        price_map[str(p.cost_model_id)].append({
            "year": p.year,
            "quarter": p.quarter,
            "price": float(p.price),
            "volume": vol.get("volume"),
            "unit": vol.get("unit"),
        })

    # Also include volume-only quarters (no price)
    for v in volumes:
        key = (str(v.cost_model_id), v.year, v.quarter)
        # Check if this quarter already added via price
        cm_id = str(v.cost_model_id)
        existing_quarters = {(r["year"], r["quarter"]) for r in price_map[cm_id]}
        if (v.year, v.quarter) not in existing_quarters:
            price_map[cm_id].append({
                "year": v.year,
                "quarter": v.quarter,
                "price": None,
                "volume": float(v.volume),
                "unit": v.unit,
            })

    models_list = []
    for m in models:
        rows = price_map.get(str(m.id), [])
        rows.sort(key=lambda r: (r["year"], r["quarter"]))
        models_list.append({
            "cost_model_id": str(m.id),
            "product_name": m.product.name if m.product else "Unknown",
            "product_reference": m.product.formula if m.product else None,
            "region": m.region,
            "currency": m.currency,
            "rows": rows,
        })
    return {
        "supplier_name": supplier.name,
        "supplier_country": supplier.country,
        "models": models_list,
    }


@router.get("/{supplier_id}/export-excel")
def export_supplier_excel(
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all cost models for a supplier as an Excel file with one sheet per model and charts."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.services.costing_engine import calculate_should_cost, _compute_indexed_cost, _apply_margin, _component_base

    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    require_team_access(db, current_user, supplier.team_id)

    models = db.query(CostModel).filter(CostModel.supplier_id == supplier_id).all()
    model_ids = [m.id for m in models]

    prices = (
        db.query(ActualPrice).filter(ActualPrice.cost_model_id.in_(model_ids))
        .order_by(ActualPrice.year, ActualPrice.quarter).all()
    ) if model_ids else []

    volumes = (
        db.query(ActualVolume).filter(ActualVolume.cost_model_id.in_(model_ids))
        .order_by(ActualVolume.year, ActualVolume.quarter).all()
    ) if model_ids else []

    # Index by (cost_model_id, year, quarter)
    price_map = defaultdict(dict)
    for p in prices:
        price_map[str(p.cost_model_id)][(p.year, p.quarter)] = float(p.price)

    vol_map = defaultdict(dict)
    for v in volumes:
        vol_map[str(v.cost_model_id)][(v.year, v.quarter)] = float(v.volume)

    # Styling
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="2D3436")
    subtitle_font = Font(name="Calibri", size=11, color="636E72")
    number_font = Font(name="Calibri", size=11)
    gap_pos_font = Font(name="Calibri", size=11, color="E74C3C")
    gap_neg_font = Font(name="Calibri", size=11, color="00B894")
    thin_border = Border(
        bottom=Side(style="thin", color="DFE6E9"),
    )
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    wb.remove(wb.active)

    for m in models:
        cm_id = str(m.id)
        product_name = m.product.name if m.product else "Unknown"
        product_ref = m.product.formula if m.product else ""
        sheet_name = f"{product_ref or product_name}"[:31]

        # Deduplicate sheet names
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            suffix = 2
            while f"{sheet_name[:28]}_{suffix}" in existing:
                suffix += 1
            sheet_name = f"{sheet_name[:28]}_{suffix}"

        ws = wb.create_sheet(title=sheet_name)

        # Collect all quarters for this model
        all_quarters = sorted(
            set(price_map[cm_id].keys()) | set(vol_map[cm_id].keys())
        )
        if not all_quarters:
            ws.append([f"No data available for {product_name}"])
            continue

        # Compute should-cost / theoretical for each quarter
        fv = m.current_formula
        sc_map = {}
        indexed_map = {}
        margin_map = {}
        if fv:
            base_price = float(fv.base_price)
            region = m.region
            for year, quarter in all_quarters:
                period_fv = m.formula_for_period(year, quarter)
                pbp = float(period_fv.base_price)
                indexed_cost = _compute_indexed_cost(
                    db, period_fv, m, region,
                    period_fv.base_year, period_fv.base_quarter,
                    year, quarter, pbp,
                )
                theoretical, margin_amt = _apply_margin(
                    indexed_cost, period_fv.margin_type, period_fv.margin_value, pbp,
                )
                sc_map[(year, quarter)] = round(theoretical, 2)
                indexed_map[(year, quarter)] = round(indexed_cost, 2)
                margin_map[(year, quarter)] = round(margin_amt, 2)

        # Title area
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{product_name}"
        ws["A1"].font = title_font
        ws.merge_cells("A2:H2")
        ws["A2"] = f"Reference: {product_ref or '—'}  |  Region: {m.region}  |  Currency: {m.currency}"
        ws["A2"].font = subtitle_font

        # Data table starting at row 4
        # Columns: Quarter | Actual Price | Should-Cost | Indexed Cost | Margin | Gap | Gap% | Volume | Revenue
        data_start_row = 4
        headers = ["Quarter", "Actual Price", "Should-Cost", "Indexed Cost", "Margin",
                    "Gap", "Gap %", "Volume", "Revenue"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=data_start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for row_idx, (year, quarter) in enumerate(all_quarters, data_start_row + 1):
            label = f"Q{quarter} {year}"
            price = price_map[cm_id].get((year, quarter))
            should_cost = sc_map.get((year, quarter))
            indexed_cost = indexed_map.get((year, quarter))
            margin_amt = margin_map.get((year, quarter))
            volume = vol_map[cm_id].get((year, quarter))
            revenue = (price * volume) if (price and volume) else None
            gap = (price - should_cost) if (price is not None and should_cost is not None) else None
            gap_pct = (gap / should_cost * 100) if (gap is not None and should_cost) else None

            ws.cell(row=row_idx, column=1, value=label).font = number_font
            ws.cell(row=row_idx, column=1).alignment = center

            c = ws.cell(row=row_idx, column=2, value=price)
            c.font = number_font; c.number_format = '#,##0'

            c = ws.cell(row=row_idx, column=3, value=should_cost)
            c.font = number_font; c.number_format = '#,##0'

            c = ws.cell(row=row_idx, column=4, value=indexed_cost)
            c.font = number_font; c.number_format = '#,##0'

            c = ws.cell(row=row_idx, column=5, value=margin_amt)
            c.font = number_font; c.number_format = '#,##0'

            c = ws.cell(row=row_idx, column=6, value=gap)
            c.font = gap_pos_font if (gap and gap > 0) else gap_neg_font if (gap and gap < 0) else number_font
            c.number_format = '#,##0'

            c = ws.cell(row=row_idx, column=7, value=round(gap_pct, 1) if gap_pct is not None else None)
            c.font = gap_pos_font if (gap_pct and gap_pct > 0) else gap_neg_font if (gap_pct and gap_pct < 0) else number_font
            c.number_format = '0.0"%"'

            c = ws.cell(row=row_idx, column=8, value=volume)
            c.font = number_font; c.number_format = '#,##0.0'

            c = ws.cell(row=row_idx, column=9, value=revenue)
            c.font = number_font; c.number_format = '#,##0'

            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).border = thin_border

        data_end_row = data_start_row + len(all_quarters)

        # Column widths
        for col, w in [("A", 12), ("B", 14), ("C", 14), ("D", 14), ("E", 12),
                       ("F", 12), ("G", 10), ("H", 12), ("I", 14)]:
            ws.column_dimensions[col].width = w

        # Categories reference (quarter labels)
        cats = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)

        # --- Price vs Should-Cost chart (line chart with both series) ---
        price_chart = LineChart()
        price_chart.title = "Actual Price vs Should-Cost"
        price_chart.y_axis.title = f"{m.currency} / unit"
        price_chart.style = 10
        price_chart.width = 24
        price_chart.height = 14

        actual_data = Reference(ws, min_col=2, min_row=data_start_row,
                                max_row=data_end_row)
        sc_data = Reference(ws, min_col=3, min_row=data_start_row,
                            max_row=data_end_row)
        price_chart.add_data(actual_data, titles_from_data=True)
        price_chart.add_data(sc_data, titles_from_data=True)
        price_chart.set_categories(cats)
        # Style: actual = blue solid, should-cost = green dashed
        price_chart.series[0].graphicalProperties.line.width = 25000
        price_chart.series[0].graphicalProperties.line.solidFill = "0984E3"
        price_chart.series[1].graphicalProperties.line.width = 25000
        price_chart.series[1].graphicalProperties.line.solidFill = "00B894"
        price_chart.series[1].graphicalProperties.line.dashStyle = "dash"
        ws.add_chart(price_chart, "K4")

        # --- Cost breakdown chart (stacked: indexed cost + margin = should-cost) ---
        breakdown_chart = BarChart()
        breakdown_chart.type = "col"
        breakdown_chart.grouping = "stacked"
        breakdown_chart.title = "Should-Cost Breakdown"
        breakdown_chart.y_axis.title = f"{m.currency} / unit"
        breakdown_chart.style = 10
        breakdown_chart.width = 24
        breakdown_chart.height = 14

        idx_data = Reference(ws, min_col=4, min_row=data_start_row,
                             max_row=data_end_row)
        margin_data = Reference(ws, min_col=5, min_row=data_start_row,
                                max_row=data_end_row)
        breakdown_chart.add_data(idx_data, titles_from_data=True)
        breakdown_chart.add_data(margin_data, titles_from_data=True)
        breakdown_chart.set_categories(cats)
        breakdown_chart.series[0].graphicalProperties.solidFill = "74B9FF"
        breakdown_chart.series[1].graphicalProperties.solidFill = "FDCB6E"
        ws.add_chart(breakdown_chart, "K20")

        # --- Volume chart ---
        vol_chart = BarChart()
        vol_chart.title = "Volume (MT)"
        vol_chart.y_axis.title = "Volume"
        vol_chart.style = 10
        vol_chart.width = 24
        vol_chart.height = 14

        vol_data = Reference(ws, min_col=8, min_row=data_start_row,
                             max_row=data_end_row)
        vol_chart.add_data(vol_data, titles_from_data=True)
        vol_chart.set_categories(cats)
        vol_chart.series[0].graphicalProperties.solidFill = "A29BFE"
        ws.add_chart(vol_chart, "K36")

    if not models:
        ws = wb.create_sheet(title="No Data")
        ws["A1"] = f"No cost models found for {supplier.name}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{supplier.name.replace(' ', '_')}_Cost_Models.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/{supplier_id}")
def delete_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    require_team_access(db, current_user, supplier.team_id)
    log_event(db, supplier.team_id, current_user.id, "delete", "supplier", str(supplier.id),
              previous_value={"name": supplier.name, "country": supplier.country})
    db.delete(supplier)
    db.commit()
    return {"status": "deleted"}
